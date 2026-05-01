import calendar
import csv
import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from django.views.generic import TemplateView
from django.db.models import Avg, Count, ExpressionWrapper, F, Q, DurationField
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from weasyprint import HTML

from tickets.models import Ticket, Status, Priority
from departments.models import Department, Category
from identity.models import User, Role
from identity.views import ManagerOrAdminRequiredMixin


# Sadece MANAGER veya ADMIN raporlara erişebilir; AGENT/EMPLOYEE engellenir.
def _require_manager_or_admin(user):
    return user.is_authenticated and user.role in (Role.MANAGER, Role.ADMIN)


# Raporlama dashboard'u — sadece MANAGER ve ADMIN erişebilir
class ReportDashboardView(ManagerOrAdminRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Tarih aralığı filtresi (?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD)
        date_from_str = self.request.GET.get('date_from', '')
        date_to_str = self.request.GET.get('date_to', '')
        date_from = _parse_date(date_from_str)
        date_to = _parse_date(date_to_str)
        context['date_from'] = date_from_str
        context['date_to'] = date_to_str

        dept_qs = Department.objects.all()
        cat_qs = Category.objects.all()
        user_qs = User.objects.filter(role__in=[Role.AGENT, Role.MANAGER])
        ticket_qs = Ticket.objects.all()

        if date_from:
            ticket_qs = ticket_qs.filter(created_at__date__gte=date_from)
        if date_to:
            ticket_qs = ticket_qs.filter(created_at__date__lte=date_to)

        if user.role == Role.MANAGER:
            dept_qs = dept_qs.filter(id=user.department_id)
            cat_qs = cat_qs.filter(department=user.department)
            user_qs = user_qs.filter(department=user.department)
            ticket_qs = ticket_qs.filter(department=user.department)

        # Departman Performans Raporları

        # Departman bazlı bilet sayıları (açık, işlemde, kapalı, toplam)
        departments = dept_qs.annotate(
            total_tickets=Count('tickets'),
            open_tickets=Count('tickets', filter=Q(tickets__status=Status.OPEN)),
            in_progress_tickets=Count('tickets', filter=Q(tickets__status=Status.IN_PROGRESS)),
            closed_tickets=Count('tickets', filter=Q(tickets__status=Status.CLOSED)),
        ).order_by('-total_tickets')

        context['departments'] = departments

        # Departman bazlı ortalama çözüm süresi — tek sorgu ile hesaplanır.
        # ticket_qs üzerinden hesaplanır → manager scope ve tarih filtresine uyar.
        dept_resolution_qs = (
            ticket_qs
            .filter(status=Status.CLOSED, closed_at__isnull=False)
            .values('department_id')
            .annotate(
                avg_duration=Avg(
                    ExpressionWrapper(F('closed_at') - F('created_at'), output_field=DurationField())
                ),
                closed_count=Count('id'),
            )
        )
        dept_resolution_map = {r['department_id']: r for r in dept_resolution_qs}

        dept_avg_resolution = []
        for dept in departments:
            res = dept_resolution_map.get(dept.pk)
            if res and res['avg_duration']:
                avg_hours = round(res['avg_duration'].total_seconds() / 3600, 1)
                closed_count = res['closed_count']
            else:
                avg_hours = None
                closed_count = 0
            dept_avg_resolution.append({
                'name': dept.name,
                'avg_hours': avg_hours,
                'closed_count': closed_count,
            })

        context['dept_avg_resolution'] = dept_avg_resolution

        # En çok talep alan kategoriler (ilk 10)
        top_categories = cat_qs.annotate(
            ticket_count=Count('tickets'),
        ).filter(ticket_count__gt=0).order_by('-ticket_count')[:10]

        context['top_categories'] = top_categories

        # Personel İş Yükü Raporları

        # Personel başına aktif bilet sayısı
        personnel = user_qs.annotate(
            active_tickets=Count(
                'assigned_tickets',
                filter=Q(assigned_tickets__status=Status.IN_PROGRESS),
            ),
            total_closed=Count(
                'assigned_tickets',
                filter=Q(assigned_tickets__status=Status.CLOSED),
            ),
        ).order_by('-active_tickets')

        context['personnel'] = personnel

        # En çok bilet üstlenen personel sıralaması (ilk 10)
        top_personnel = user_qs.annotate(
            total_assigned=Count('assigned_tickets'),
        ).filter(total_assigned__gt=0).order_by('-total_assigned')[:10]

        context['top_personnel'] = top_personnel

        # Personel bazlı ortalama çözüm süresi — ticket_qs scope'una uyar
        personnel_resolution_qs = (
            ticket_qs
            .filter(status=Status.CLOSED, closed_at__isnull=False, assigned_to__isnull=False)
            .values('assigned_to_id')
            .annotate(
                avg_duration=Avg(
                    ExpressionWrapper(F('closed_at') - F('created_at'), output_field=DurationField())
                ),
            )
        )
        personnel_resolution_map = {r['assigned_to_id']: r for r in personnel_resolution_qs}

        personnel_avg = []
        for p in top_personnel:
            res = personnel_resolution_map.get(p.pk)
            if res and res['avg_duration']:
                avg_hours = round(res['avg_duration'].total_seconds() / 3600, 1)
            else:
                avg_hours = None
            personnel_avg.append({
                'name': p.get_full_name() or p.username,
                'avg_hours': avg_hours,
            })

        context['personnel_avg'] = personnel_avg

        # Grafiksel Dashboard Verileri

        # Genel bilet istatistikleri
        context['total_tickets'] = ticket_qs.count()
        context['open_count'] = ticket_qs.filter(status=Status.OPEN).count()
        context['in_progress_count'] = ticket_qs.filter(status=Status.IN_PROGRESS).count()
        context['closed_count'] = ticket_qs.filter(status=Status.CLOSED).count()

        # Aylık bilet trend verisi (son 6 ay) — takvim tabanlı doğru ay hesabı
        context['monthly_labels'], context['monthly_counts'] = _get_monthly_trend(ticket_qs)

        # Departman karşılaştırma verisi (çubuk grafik için)
        context['dept_names'] = [d.name for d in departments]
        context['dept_open'] = [d.open_tickets for d in departments]
        context['dept_in_progress'] = [d.in_progress_tickets for d in departments]
        context['dept_closed'] = [d.closed_tickets for d in departments]

        return context


# Ortak yardımcı fonksiyonlar

def _parse_date(date_str):
    """YYYY-MM-DD formatındaki string'i date nesnesine dönüştürür."""
    if not date_str:
        return None
    try:
        return date.fromisoformat(date_str)
    except ValueError:
        return None


def _get_monthly_trend(ticket_qs):
    """Son 6 ayın bilet sayılarını gerçek takvim aylarına göre hesaplar."""
    today = timezone.now()
    labels = []
    counts = []
    for i in range(5, -1, -1):
        total_months = today.year * 12 + today.month - 1 - i
        year = total_months // 12
        month = total_months % 12 + 1
        _, last_day = calendar.monthrange(year, month)
        month_start = today.replace(year=year, month=month, day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = today.replace(year=year, month=month, day=last_day, hour=23, minute=59, second=59, microsecond=0)
        count = ticket_qs.filter(created_at__gte=month_start, created_at__lte=month_end).count()
        labels.append(month_start.strftime('%b %Y'))
        counts.append(count)
    return labels, counts


def _get_ticket_export_data(user, date_from=None, date_to=None):
    """Dışa aktarım için bilet verilerini toplar (rol bazlı kapsamla)."""
    qs = Ticket.objects.select_related(
        'sender', 'assigned_to', 'department', 'category',
    ).order_by('-created_at')

    # Rol bazlı kapsam: ADMIN tümü, MANAGER kendi departmanı; AGENT/EMPLOYEE bu noktaya
    # ulaşamamalı (view-seviyesinde 403 döner) ama defansif filtre eklenir.
    if user.role == Role.MANAGER:
        qs = qs.filter(department=user.department)
    elif user.role != Role.ADMIN:
        qs = qs.none()

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    rows = []
    for t in qs:
        rows.append({
            'id': t.pk,
            'subject': t.subject,
            'status': t.get_status_display(),
            'priority': t.get_priority_display(),
            'department': t.department.name if t.department else '—',
            'category': t.category.name if t.category else '—',
            'sender': (t.sender.get_full_name() or t.sender.username) if t.sender else '—',
            'assigned_to': (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to else '—',
            'created_at': t.created_at.strftime('%d.%m.%Y %H:%M'),
            'closed_at': t.closed_at.strftime('%d.%m.%Y %H:%M') if t.closed_at else '—',
            'resolution_note': t.resolution_note or '',
        })
    return rows


EXPORT_HEADERS = [
    'ID', 'Konu', 'Durum', 'Öncelik', 'Departman', 'Kategori',
    'Talep Sahibi', 'Üstlenen Personel', 'Oluşturulma', 'Kapatılma', 'Çözüm Notu',
]


# CSV Dışa Aktarım — Sadece MANAGER/ADMIN

@login_required
def export_csv(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    date_from = _parse_date(request.GET.get('date_from', ''))
    date_to = _parse_date(request.GET.get('date_to', ''))
    rows = _get_ticket_export_data(request.user, date_from, date_to)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.csv"'
    response.write('﻿')  # BOM for Excel UTF-8 compatibility

    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)
    for r in rows:
        writer.writerow([
            r['id'], r['subject'], r['status'], r['priority'],
            r['department'], r['category'], r['sender'], r['assigned_to'],
            r['created_at'], r['closed_at'], r['resolution_note'],
        ])
    return response


# Excel Dışa Aktarım — Sadece MANAGER/ADMIN

@login_required
def export_excel(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    date_from = _parse_date(request.GET.get('date_from', ''))
    date_to = _parse_date(request.GET.get('date_to', ''))
    rows = _get_ticket_export_data(request.user, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bilet Raporu'

    # Başlık satırı stili
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')

    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Veri satırları
    keys = [
        'id', 'subject', 'status', 'priority', 'department', 'category',
        'sender', 'assigned_to', 'created_at', 'closed_at', 'resolution_note',
    ]
    for row_idx, r in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=r[key])

    # Sütun genişlikleri
    col_widths = [6, 30, 10, 10, 18, 18, 20, 20, 18, 18, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.xlsx"'
    return response


# PDF Dışa Aktarım — Sadece MANAGER/ADMIN

@login_required
def export_pdf(request):
    if not _require_manager_or_admin(request.user):
        return HttpResponseForbidden('Bu rapora erişim yetkiniz bulunmamaktadır.')
    date_from = _parse_date(request.GET.get('date_from', ''))
    date_to = _parse_date(request.GET.get('date_to', ''))
    rows = _get_ticket_export_data(request.user, date_from, date_to)

    # İstatistikler
    total = len(rows)
    open_count = sum(1 for r in rows if r['status'] == 'Açık')
    in_progress_count = sum(1 for r in rows if r['status'] == 'İşlemde')
    closed_count = sum(1 for r in rows if r['status'] == 'Kapalı')

    html = render_to_string('reports/export_pdf.html', {
        'rows': rows,
        'total': total,
        'open_count': open_count,
        'in_progress_count': in_progress_count,
        'closed_count': closed_count,
        'generated_at': timezone.now().strftime('%d.%m.%Y %H:%M'),
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bilet_raporu.pdf"'

    HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(response)
    return response
